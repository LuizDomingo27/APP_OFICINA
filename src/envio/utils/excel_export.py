"""Exportação Excel com layout executivo e formatação padrão BR."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from src.envio.config.settings import HIDDEN_COLUMNS

INK = "0F1B2D"
ACCENT = "0E9F6E"
NEON_SOFT = "E8FBF2"
BORDER = "E6ECEF"
WHITE = "FFFFFF"

BR_DATE = "dd/mm/yyyy"
BR_INTEGER = "#.##0"

HEADER_LABELS = {
    "ENVIO": "Data Envio",
    "QTD": "Qtd. Peças",
    "MINUTOS": "Minutos",
    "MP": "Matéria-Prima",
    "PDV": "PDV",
    "OFICINA": "Oficina",
    "FRETE": "Frete",
    "ORIGEM": "Origem",
    "SITUAÇÃO": "Situação",
}

NUMERIC_COLUMNS = frozenset({"QTD", "MINUTOS"})
DATE_COLUMNS = frozenset({"ENVIO"})


def _thin_border() -> Border:
    side = Side(style="thin", color=BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _column_format(col_name: str) -> str | None:
    if col_name in DATE_COLUMNS:
        return BR_DATE
    if col_name in NUMERIC_COLUMNS:
        return BR_INTEGER
    return None


def _export_columns(df: pd.DataFrame) -> list[str]:
    columns = [c for c in HEADER_LABELS if c in df.columns and c not in HIDDEN_COLUMNS]
    columns.extend(
        c for c in df.columns if c not in columns and c not in HIDDEN_COLUMNS
    )
    return columns or df.columns.tolist()


def _prepare_export_frame(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    export_df = df[columns].copy()
    if "ENVIO" in export_df.columns:
        export_df["ENVIO"] = pd.to_datetime(export_df["ENVIO"], errors="coerce")
    for col_name in NUMERIC_COLUMNS:
        if col_name in export_df.columns:
            export_df[col_name] = pd.to_numeric(export_df[col_name], errors="coerce").fillna(0)
    return export_df


def build_executive_excel(
    df: pd.DataFrame,
    date_start: date,
    date_end: date,
) -> bytes:
    """Gera planilha .xlsx estilizada para leitura executiva."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Envios"

    columns = _export_columns(df)
    last_col = len(columns)
    last_col_letter = get_column_letter(last_col)
    header_row = 5
    data_start = header_row + 1

    title_font = Font(name="Calibri", size=16, bold=True, color=INK)
    subtitle_font = Font(name="Calibri", size=11, color="6B7A90")
    meta_font = Font(name="Calibri", size=10, color="6B7A90", italic=True)
    header_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    total_font = Font(name="Calibri", size=10, bold=True, color=INK)

    header_fill = PatternFill("solid", fgColor=INK)
    total_fill = PatternFill("solid", fgColor=NEON_SOFT)
    border = _thin_border()

    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "ENVIOS PARA OFICINAS"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = (
        f"Relatório Executivo · Período: "
        f"{date_start.strftime('%d/%m/%Y')} a {date_end.strftime('%d/%m/%Y')}"
    )
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"A3:{last_col_letter}3")
    ws["A3"] = (
        f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} · "
        f"{len(df):,} registro(s)".replace(",", ".")
    )
    ws["A3"].font = meta_font
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[header_row].height = 22

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=HEADER_LABELS.get(col_name, col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    export_df = _prepare_export_frame(df, columns)
    for row_values in dataframe_to_rows(export_df, index=False, header=False):
        ws.append(row_values)

    data_end = data_start + len(export_df) - 1
    if len(export_df) > 0:
        for col_idx, col_name in enumerate(columns, start=1):
            fmt = _column_format(col_name)
            if not fmt:
                continue
            for excel_row in range(data_start, data_end + 1):
                ws.cell(row=excel_row, column=col_idx).number_format = fmt

    total_row = data_start + len(export_df)
    if len(export_df) > 0:
        ws.row_dimensions[total_row].height = 22
        label_col = columns.index("ENVIO") + 1 if "ENVIO" in columns else 1

        total_border = Border(
            top=Side(style="medium", color=ACCENT),
            left=Side(style="thin", color=BORDER),
            right=Side(style="thin", color=BORDER),
            bottom=Side(style="thin", color=BORDER),
        )
        for col_idx in range(1, last_col + 1):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.fill = total_fill
            cell.border = total_border

        label_cell = ws.cell(row=total_row, column=label_col, value="TOTAL")
        label_cell.font = total_font
        label_cell.alignment = Alignment(horizontal="right", vertical="center")

        for col_name in NUMERIC_COLUMNS:
            if col_name not in columns:
                continue
            col_idx = columns.index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            total_cell = ws.cell(
                row=total_row,
                column=col_idx,
                value=f"=SUM({col_letter}{data_start}:{col_letter}{total_row - 1})",
            )
            total_cell.font = total_font
            total_cell.alignment = Alignment(horizontal="right", vertical="center")
            total_cell.number_format = BR_INTEGER

        last_data_row = total_row
    else:
        last_data_row = header_row

    for col_idx, col_name in enumerate(columns, start=1):
        header_len = len(HEADER_LABELS.get(col_name, col_name))
        sample = export_df[col_name].astype(str).head(200).map(len).max() if len(export_df) else 0
        width = max(header_len, sample or 0, 12)
        if col_name in NUMERIC_COLUMNS:
            width = max(width, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 42)

    ws.freeze_panes = ws.cell(row=data_start, column=1).coordinate
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_data_row}"

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"{header_row}:{header_row}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@st.cache_data(show_spinner=False)
def build_executive_excel_cached(
    export_key: str,
    df: pd.DataFrame,
    date_start: date,
    date_end: date,
) -> bytes:
    """Cacheia o Excel exportado por combinação de filtros e dados."""
    return build_executive_excel(df, date_start, date_end)
