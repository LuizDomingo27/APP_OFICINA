"""
page.py — Painel de Relatório Executivo Consolidadado.
Este painel combina os dados de Envios e Recebimentos para apresentar métricas executivas,
como saldo pendente em oficinas, taxa de retorno, e gráficos comparativos de desempenho.
"""

from typing import Any
from datetime import date
from io import BytesIO
import pandas as pd
import streamlit as st
from src.shared.echarts_utils import safe_st_echarts
from src.shared.state import resetar_sessao

# Configuração de Cores Executivas
COLOR_PRIMARY = "#0E9F6E"  # Verde neon-aqua (padrão visual do Painel de Recebimento)
COLOR_NEON = "#0C8560"     # Verde de apoio/gradiente (idem Recebimento)
COLOR_DARK = "#0F1B2D"     # Ink/Escuro
COLOR_MUTED = "#6B7A90"    # Muted
COLOR_BORDER = "#E6ECEF"   # Borda
COLOR_CARD = "#FFFFFF"     # Card

def _inject_css() -> None:
    st.markdown(
        f"""
        <style>
        html, body, [class*="css"] {{
            font-family: 'Inter', sans-serif;
            color: {COLOR_DARK} !important;
        }}
        .metric-card-exec {{
            background: {COLOR_CARD};
            border: 1px solid {COLOR_BORDER};
            border-left: 6px solid {COLOR_PRIMARY};
            border-radius: 12px;
            padding: 16px 20px;
            box-shadow: 0 4px 12px rgba(15,27,45,0.03);
            text-align: left;
        }}
        .metric-card-exec .label {{
            color: {COLOR_MUTED};
            font-size: 0.75rem;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.07em;
        }}
        .metric-card-exec .value {{
            font-size: 1.8rem;
            font-weight: 800;
            color: {COLOR_DARK};
            margin-top: 4px;
        }}
        .metric-card-exec .sub {{
            font-size: 0.8rem;
            color: {COLOR_MUTED};
            margin-top: 4px;
        }}
        .executive-table-wrapper {{
            background: {COLOR_CARD};
            border: 1px solid {COLOR_BORDER};
            border-radius: 12px;
            box-shadow: 0 4px 12px rgba(15,27,45,0.03);
            overflow: hidden;
            margin-bottom: 20px;
        }}
        .executive-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 0.88rem;
        }}
        .executive-table th {{
            background: #0E9F6E !important;
            color: #FFFFFF !important;
            font-weight: 700;
            text-transform: uppercase;
            font-size: 0.68rem;
            letter-spacing: 0.07em;
            padding: 13px 14px;
            border-bottom: 2px solid #0C8560;
            text-align: left;
        }}
        .executive-table th.num {{
            text-align: center;
        }}
        .executive-table td {{
            padding: 12px 16px;
            border-bottom: 1px solid #F0F4F6;
            color: {COLOR_DARK};
            text-align: left;
        }}
        .executive-table td.num {{
            text-align: center;
            font-variant-numeric: tabular-nums;
        }}
        .executive-table tr:hover {{
            background-color: #FAFCFB;
        }}
        </style>
        """,
        unsafe_allow_html=True
    )

def _build_kpis(df_envio: pd.DataFrame, df_receb: pd.DataFrame) -> dict:
    total_enviadas = df_envio["QTD"].sum()
    total_recebidas = df_receb["REAL CORTADO"].sum()
    total_minutos_enviados = df_envio["MINUTOS"].sum()
    total_minutos_recebidos = df_receb["MINUTOS"].sum()

    return {
        "enviadas": total_enviadas,
        "recebidas": total_recebidas,
        "minutos_enviados": total_minutos_enviados,
        "minutos_recebidos": total_minutos_recebidos,
    }

import unicodedata

def _normalize_oficina(name: Any) -> str:
    if pd.isna(name):
        return ""
    name_str = str(name).strip().upper()
    # Remove acentos
    name_str = "".join(
        c for c in unicodedata.normalize("NFD", name_str)
        if unicodedata.category(c) != "Mn"
    )
    # Remove espaços duplos
    return " ".join(name_str.split())

@st.cache_data(show_spinner=False)
def _build_workshop_summary(df_envio: pd.DataFrame, df_receb: pd.DataFrame) -> pd.DataFrame:
    df_env = df_envio.copy()
    df_rec = df_receb.copy()
    
    # Otimização: calcula normalização apenas para valores únicos
    unique_env = df_env["OFICINA"].dropna().unique()
    env_map = {name: _normalize_oficina(name) for name in unique_env}
    df_env["OFICINA_NORM"] = df_env["OFICINA"].map(env_map).fillna("")
    
    unique_rec = df_rec["OFICINA"].dropna().unique()
    rec_map = {name: _normalize_oficina(name) for name in unique_rec}
    df_rec["OFICINA_NORM"] = df_rec["OFICINA"].map(rec_map).fillna("")
    
    # Agrupa Envio por Oficina Normalizada
    envio_g = df_env.groupby("OFICINA_NORM", as_index=False)["QTD"].sum().rename(columns={"QTD": "ENVIADAS", "OFICINA_NORM": "OFICINA"})
    # Agrupa Recebimento por Oficina Normalizada
    receb_g = df_rec.groupby("OFICINA_NORM", as_index=False)["REAL CORTADO"].sum().rename(columns={"REAL CORTADO": "RECEBIDAS", "OFICINA_NORM": "OFICINA"})
    
    # Merge
    merged = pd.merge(envio_g, receb_g, on="OFICINA", how="outer").fillna(0)
    merged["ENVIADAS"] = merged["ENVIADAS"].astype(int)
    merged["RECEBIDAS"] = merged["RECEBIDAS"].astype(int)
    merged["SALDO_PENDENTE"] = merged["ENVIADAS"] - merged["RECEBIDAS"]
    merged["TAXA_RETORNO"] = (merged["RECEBIDAS"] / merged["ENVIADAS"] * 100).where(merged["ENVIADAS"] > 0, 0.0)
    
    return merged.sort_values(by="ENVIADAS", ascending=False).reset_index(drop=True)


from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

def _build_chart_comparison(merged_daily: pd.DataFrame) -> dict:
    dates = []
    for d in merged_daily["DIA"]:
        try:
            dates.append(pd.Timestamp(d).strftime("%d/%m"))
        except Exception:
            dates.append(str(d))
            
    enviadas = [int(v) for v in merged_daily["ENVIADAS"]]
    recebidas = [int(v) for v in merged_daily["RECEBIDAS"]]
    
    from src.envio.charts.theme import tooltip_dark, grid, axis_label, split_line
    from src.shared.echarts_utils import JsCode
    
    tip_fmt = JsCode("""
    function(params) {
        var out = '<div style="font-weight:600;margin-bottom:5px">' + params[0].axisValueLabel + '</div>';
        params.forEach(function(p) {
            out += p.marker + ' ' + p.seriesName +
                ': <b>' + p.value.toLocaleString('pt-BR') + '</b><br/>';
        });
        return out;
    }
    """)
    
    _FMT_K = JsCode(
        "function(v){"
        "  if(v>=1000000) return (v/1000000).toFixed(1).replace('.',',')+' M';"
        "  if(v>=1000)    return (v/1000).toFixed(1).replace('.',',')+' k';"
        "  return v;"
        "}"
    )
    
    area_envio = {
        "type": "linear", "x": 0, "y": 0, "x2": 0, "y2": 1,
        "colorStops": [
            {"offset": 0, "color": "rgba(107,122,144,0.15)"},
            {"offset": 1, "color": "rgba(107,122,144,0.01)"},
        ],
    }
    
    area_recebido = {
        "type": "linear", "x": 0, "y": 0, "x2": 0, "y2": 1,
        "colorStops": [
            {"offset": 0, "color": "rgba(14,159,110,0.18)"},
            {"offset": 1, "color": "rgba(14,159,110,0.02)"},
        ],
    }
    
    return {
        "backgroundColor": "transparent",
        "color": ["#6B7A90", "#0E9F6E"],
        "tooltip": {
            "trigger": "axis",
            "axisPointer": {
                "type": "cross",
                "label": {"backgroundColor": "#334155"},
            },
            "formatter": tip_fmt,
            **tooltip_dark(),
        },
        "legend": {
            "data": ["Peças Enviadas", "Peças Recebidas"],
            "top": 2,
            "icon": "roundRect",
            "itemWidth": 16,
            "itemHeight": 8,
            "itemGap": 20,
            "textStyle": {
                "fontSize": 12,
                "color": "#475569",
                "fontFamily": "Inter, sans-serif",
            },
        },
        "grid": grid(bottom="12%", top="16%"),
        "xAxis": {
            "type": "category",
            "data": dates,
            "boundaryGap": False,
            "axisLabel": axis_label(rotate=30),
            "axisLine": {"lineStyle": {"color": "#E2E8F0"}},
            "axisTick": {"show": False},
        },
        "yAxis": {
            "type": "value",
            "name": "Peças",
            "nameTextStyle": {"color": "#94A3B8", "fontSize": 12},
            "axisLabel": {**axis_label(), "formatter": _FMT_K},
            "splitLine": split_line(),
        },
        "series": [
            {
                "name": "Peças Enviadas",
                "type": "line",
                "smooth": True,
                "data": enviadas,
                "symbol": "circle",
                "symbolSize": 5,
                "itemStyle": {
                    "color": "#6B7A90",
                    "borderColor": "#FFFFFF",
                    "borderWidth": 1.5,
                },
                "lineStyle": {"width": 2, "color": "#6B7A90", "type": "dashed"},
                "areaStyle": {"color": area_envio},
            },
            {
                "name": "Peças Recebidas",
                "type": "line",
                "smooth": True,
                "data": recebidas,
                "symbol": "circle",
                "symbolSize": 6,
                "itemStyle": {
                    "color": "#0E9F6E",
                    "borderColor": "#FFFFFF",
                    "borderWidth": 2,
                },
                "lineStyle": {"width": 2.5, "color": "#0E9F6E"},
                "areaStyle": {"color": area_recebido},
            },
        ],
    }

@st.cache_data(show_spinner=False)
def _to_excel(summary_df: pd.DataFrame, df_envio: pd.DataFrame, df_receb: pd.DataFrame, date_start: date, date_end: date) -> bytes:
    wb = Workbook()
    
    INK = "0F1B2D"
    ACCENT = "0E9F6E"
    NEON_SOFT = "E8FBF2"
    BORDER = "E6ECEF"
    WHITE = "FFFFFF"
    
    title_font = Font(name="Calibri", size=16, bold=True, color=INK)
    subtitle_font = Font(name="Calibri", size=11, color="6B7A90")
    meta_font = Font(name="Calibri", size=10, color="6B7A90", italic=True)
    header_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    total_font = Font(name="Calibri", size=10, bold=True, color=INK)
    
    header_fill = PatternFill("solid", fgColor=INK)
    total_fill = PatternFill("solid", fgColor=NEON_SOFT)
    
    side = Side(style="thin", color=BORDER)
    thin_border = Border(left=side, right=side, top=side, bottom=side)
    
    # --- PLANILHA 1: Consolidado Oficinas ---
    ws_summary = wb.active
    ws_summary.title = "Consolidado Oficinas"
    ws_summary.sheet_view.showGridLines = False
    
    # Título da planilha 1
    ws_summary.merge_cells("A1:E1")
    ws_summary["A1"] = "CONSOLIDADO DE OFICINAS"
    ws_summary["A1"].font = title_font
    ws_summary["A1"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws_summary.merge_cells("A2:E2")
    ws_summary["A2"] = f"Relatório Consolidado · Período: {date_start.strftime('%d/%m/%Y')} a {date_end.strftime('%d/%m/%Y')}"
    ws_summary["A2"].font = subtitle_font
    ws_summary["A2"].alignment = Alignment(horizontal="left", vertical="center")
    
    from datetime import datetime
    ws_summary.merge_cells("A3:E3")
    ws_summary["A3"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_summary["A3"].font = meta_font
    ws_summary["A3"].alignment = Alignment(horizontal="left", vertical="center")
    
    # Cabeçalho da tabela
    headers = ["Oficina", "Peças Enviadas", "Peças Recebidas", "Saldo Pendente", "Taxa de Retorno"]
    header_row = 5
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_summary.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        if col_idx == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif col_idx in [2, 3, 4, 5]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    ws_summary.row_dimensions[1].height = 28
    ws_summary.row_dimensions[2].height = 20
    ws_summary.row_dimensions[3].height = 18
    ws_summary.row_dimensions[header_row].height = 22
    
    data_start = header_row + 1
    for row_idx, row in summary_df.iterrows():
        excel_row = data_start + row_idx
        ws_summary.cell(row=excel_row, column=1, value=row["OFICINA"]).alignment = Alignment(horizontal="left", vertical="center")
        ws_summary.cell(row=excel_row, column=2, value=int(row["ENVIADAS"])).number_format = "#,##0"
        ws_summary.cell(row=excel_row, column=3, value=int(row["RECEBIDAS"])).number_format = "#,##0"
        ws_summary.cell(row=excel_row, column=4, value=int(row["SALDO_PENDENTE"])).number_format = "#,##0"
        ws_summary.cell(row=excel_row, column=5, value=float(row["TAXA_RETORNO"] / 100)).number_format = "0.0%"
        
        for col_idx in range(1, 6):
            c = ws_summary.cell(row=excel_row, column=col_idx)
            c.font = Font(name="Calibri", size=10)
            c.border = thin_border
            if col_idx > 1:
                c.alignment = Alignment(horizontal="right", vertical="center")
                
    # Total row
    total_row = data_start + len(summary_df)
    ws_summary.row_dimensions[total_row].height = 22
    total_border = Border(
        top=Side(style="medium", color=ACCENT),
        left=Side(style="thin", color=BORDER),
        right=Side(style="thin", color=BORDER),
        bottom=Side(style="thin", color=BORDER),
    )
    for col_idx in range(1, 6):
        cell = ws_summary.cell(row=total_row, column=col_idx)
        cell.fill = total_fill
        cell.border = total_border
        
    ws_summary.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws_summary.cell(row=total_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    
    for col_idx, col_letter in enumerate(["B", "C", "D"], start=2):
        cell = ws_summary.cell(row=total_row, column=col_idx, value=f"=SUM({col_letter}{data_start}:{col_letter}{total_row - 1})")
        cell.font = total_font
        cell.number_format = "#,##0"
        cell.alignment = Alignment(horizontal="right", vertical="center")
        
    # Total taxa de retorno formula: Total Recebidas / Total Enviadas
    cell_taxa = ws_summary.cell(row=total_row, column=5, value=f'=IFERROR(C{total_row}/B{total_row}, 0)')
    cell_taxa.font = total_font
    cell_taxa.number_format = "0.0%"
    cell_taxa.alignment = Alignment(horizontal="right", vertical="center")
    
    # Auto-adjust column width for summary
    for col in range(1, 6):
        col_letter = get_column_letter(col)
        # Find maximum length of values in this column
        max_len = len(headers[col-1])
        for r in range(data_start, total_row + 1):
            val = ws_summary.cell(row=r, column=col).value
            if val is not None:
                # If formula, estimate length
                if str(val).startswith("="):
                    max_len = max(max_len, 12)
                else:
                    max_len = max(max_len, len(str(val)))
        ws_summary.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 40)
        
    ws_summary.freeze_panes = "A6"
    ws_summary.auto_filter.ref = f"A5:E{total_row}"
    
    # --- PLANILHA 2: Detalhe Envios ---
    ws_envio = wb.create_sheet(title="Detalhe Envios")
    ws_envio.sheet_view.showGridLines = False
    
    ws_envio.cell(row=1, column=1, value="DETALHAMENTO DE ENVIOS").font = title_font
    ws_envio.cell(row=2, column=1, value="Valores expressos no padrão monetário / numérico BR").font = subtitle_font
    
    env_cols = ["DATA", "OFICINA", "MP", "QTD", "MINUTOS"]
    ws_envio.row_dimensions[4].height = 22
    for col_idx, col_name in enumerate(env_cols, start=1):
        cell = ws_envio.cell(row=4, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        if col_idx in [4, 5]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif col_idx == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        
    for r_idx, row in df_envio.reset_index(drop=True).iterrows():
        excel_row = 5 + r_idx
        # Write actual date
        dt_val = pd.to_datetime(row["ENVIO"])
        cell_date = ws_envio.cell(row=excel_row, column=1, value=dt_val.to_pydatetime() if hasattr(dt_val, "to_pydatetime") else dt_val)
        cell_date.number_format = "dd/mm/yyyy"
        cell_date.alignment = Alignment(horizontal="center", vertical="center")
        
        ws_envio.cell(row=excel_row, column=2, value=row["OFICINA"]).alignment = Alignment(horizontal="left", vertical="center")
        ws_envio.cell(row=excel_row, column=3, value=row["MP"]).alignment = Alignment(horizontal="center", vertical="center")
        
        ws_envio.cell(row=excel_row, column=4, value=int(row["QTD"])).number_format = "#,##0"
        ws_envio.cell(row=excel_row, column=5, value=int(row["MINUTOS"])).number_format = "#,##0"
        
        for col_idx in range(1, 6):
            c = ws_envio.cell(row=excel_row, column=col_idx)
            c.font = Font(name="Calibri", size=10)
            c.border = thin_border
            if col_idx in [4, 5]:
                c.alignment = Alignment(horizontal="right", vertical="center")
                
    total_row_env = 5 + len(df_envio)
    ws_envio.row_dimensions[total_row_env].height = 22
    for col_idx in range(1, 6):
        cell = ws_envio.cell(row=total_row_env, column=col_idx)
        cell.fill = total_fill
        cell.border = total_border
        
    ws_envio.cell(row=total_row_env, column=1, value="TOTAL").font = total_font
    ws_envio.cell(row=total_row_env, column=1).alignment = Alignment(horizontal="left", vertical="center")
    
    for col_idx, col_letter in enumerate(["D", "E"], start=4):
        cell = ws_envio.cell(row=total_row_env, column=col_idx, value=f"=SUM({col_letter}5:{col_letter}{total_row_env - 1})")
        cell.font = total_font
        cell.number_format = "#,##0"
        cell.alignment = Alignment(horizontal="right", vertical="center")
        
    for col in range(1, 6):
        col_letter = get_column_letter(col)
        max_len = len(env_cols[col-1])
        for r in range(5, total_row_env + 1):
            val = ws_envio.cell(row=r, column=col).value
            if val is not None:
                if str(val).startswith("="):
                    max_len = max(max_len, 12)
                elif isinstance(val, (datetime, date)):
                    max_len = max(max_len, 10)
                else:
                    max_len = max(max_len, len(str(val)))
        ws_envio.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 40)
        
    ws_envio.freeze_panes = "A5"
    ws_envio.auto_filter.ref = f"A4:E{total_row_env}"

    # --- PLANILHA 3: Detalhe Recebimentos ---
    ws_receb = wb.create_sheet(title="Detalhe Recebimentos")
    ws_receb.sheet_view.showGridLines = False
    
    ws_receb.cell(row=1, column=1, value="DETALHAMENTO DE RECEBIMENTOS").font = title_font
    ws_receb.cell(row=2, column=1, value="Valores expressos no padrão monetário / numérico BR").font = subtitle_font
    
    rec_cols = ["DATA", "OFICINA", "MP", "PEÇAS RECEBIDAS", "MINUTOS"]
    ws_receb.row_dimensions[4].height = 22
    for col_idx, col_name in enumerate(rec_cols, start=1):
        cell = ws_receb.cell(row=4, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        if col_idx in [4, 5]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif col_idx == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        
    for r_idx, row in df_receb.reset_index(drop=True).iterrows():
        excel_row = 5 + r_idx
        # Write actual date
        dt_val = pd.to_datetime(row["DIA"])
        cell_date = ws_receb.cell(row=excel_row, column=1, value=dt_val.to_pydatetime() if hasattr(dt_val, "to_pydatetime") else dt_val)
        cell_date.number_format = "dd/mm/yyyy"
        cell_date.alignment = Alignment(horizontal="center", vertical="center")
        
        ws_receb.cell(row=excel_row, column=2, value=row["OFICINA"]).alignment = Alignment(horizontal="left", vertical="center")
        ws_receb.cell(row=excel_row, column=3, value=row["MP"]).alignment = Alignment(horizontal="center", vertical="center")
        
        ws_receb.cell(row=excel_row, column=4, value=int(row["REAL CORTADO"])).number_format = "#,##0"
        ws_receb.cell(row=excel_row, column=5, value=int(row["MINUTOS"])).number_format = "#,##0"
        
        for col_idx in range(1, 6):
            c = ws_receb.cell(row=excel_row, column=col_idx)
            c.font = Font(name="Calibri", size=10)
            c.border = thin_border
            if col_idx in [4, 5]:
                c.alignment = Alignment(horizontal="right", vertical="center")
                
    total_row_rec = 5 + len(df_receb)
    ws_receb.row_dimensions[total_row_rec].height = 22
    for col_idx in range(1, 6):
        cell = ws_receb.cell(row=total_row_rec, column=col_idx)
        cell.fill = total_fill
        cell.border = total_border
        
    ws_receb.cell(row=total_row_rec, column=1, value="TOTAL").font = total_font
    ws_receb.cell(row=total_row_rec, column=1).alignment = Alignment(horizontal="left", vertical="center")
    
    for col_idx, col_letter in enumerate(["D", "E"], start=4):
        cell = ws_receb.cell(row=total_row_rec, column=col_idx, value=f"=SUM({col_letter}5:{col_letter}{total_row_rec - 1})")
        cell.font = total_font
        cell.number_format = "#,##0"
        cell.alignment = Alignment(horizontal="right", vertical="center")
        
    for col in range(1, 6):
        col_letter = get_column_letter(col)
        max_len = len(rec_cols[col-1])
        for r in range(5, total_row_rec + 1):
            val = ws_receb.cell(row=r, column=col).value
            if val is not None:
                if str(val).startswith("="):
                    max_len = max(max_len, 12)
                elif isinstance(val, (datetime, date)):
                    max_len = max(max_len, 10)
                else:
                    max_len = max(max_len, len(str(val)))
        ws_receb.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 40)
        
    ws_receb.freeze_panes = "A5"
    ws_receb.auto_filter.ref = f"A4:E{total_row_rec}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

def render_relatorios_page(envio_dataset: Any, recebimento_df: pd.DataFrame) -> None:
    _inject_css()
    
    st.title("📊 Relatório Executivo Consolidado")
    st.markdown("Visão executiva integrada: cruzamento de dados de envio e recebimento de peças e tempo (minutos) por oficina.")
    
    # Extrai o df de envios
    df_envio_full = envio_dataset.df
    df_receb_full = recebimento_df
    
    # Sidebar Filtros da Página de Relatório
    with st.sidebar:
        st.markdown("### 🔍 Filtros do Relatório")
        
        # Determina datas limites comuns
        min_date = min(df_envio_full["ENVIO"].min().date(), df_receb_full["DIA"].min().date())
        max_date = max(df_envio_full["ENVIO"].max().date(), df_receb_full["DIA"].max().date())
        
        date_range = st.date_input(
            "Período de Análise",
            value=(min_date, max_date),
            min_value=min_date,
            max_value=max_date,
            format="DD/MM/YYYY"
        )
        
        # Opções de Matéria-Prima (Unificado)
        all_mps = sorted(list(set(df_envio_full["MP"].dropna().unique()).union(set(df_receb_full["MP"].dropna().unique()))))
        sel_mps = st.multiselect("Matéria-Prima", options=all_mps, placeholder="Todas")
        
        # Opções de Oficina (Unificado)
        all_oficinas = sorted(list(set(df_envio_full["OFICINA"].dropna().unique()).union(set(df_receb_full["OFICINA"].dropna().unique()))))
        sel_oficinas = st.multiselect("Oficinas", options=all_oficinas, placeholder="Todas")

    # Filtra os DataFrames com base na seleção
    if isinstance(date_range, tuple) and len(date_range) == 2:
        d_ini, d_fim = date_range
    else:
        d_ini, d_fim = min_date, max_date
        
    # Filtro Envio
    mask_env = (df_envio_full["ENVIO"].dt.date >= d_ini) & (df_envio_full["ENVIO"].dt.date <= d_fim)
    if sel_mps:
        mask_env &= df_envio_full["MP"].isin(sel_mps)
    if sel_oficinas:
        mask_env &= df_envio_full["OFICINA"].isin(sel_oficinas)
    df_envio_filtered = df_envio_full[mask_env]
    
    # Filtro Recebimento
    mask_rec = (df_receb_full["DIA"].dt.date >= d_ini) & (df_receb_full["DIA"].dt.date <= d_fim)
    if sel_mps:
        mask_rec &= df_receb_full["MP"].isin(sel_mps)
    if sel_oficinas:
        mask_rec &= df_receb_full["OFICINA"].isin(sel_oficinas)
    df_receb_filtered = df_receb_full[mask_rec]

    # Cálculos
    kpis = _build_kpis(df_envio_filtered, df_receb_filtered)
    summary_df = _build_workshop_summary(df_envio_filtered, df_receb_filtered)
    
    # Agrupamento Temporal Diário para o Gráfico
    df_envio_daily = df_envio_filtered.copy()
    df_envio_daily["DIA"] = pd.to_datetime(df_envio_daily["ENVIO"]).dt.date
    envio_daily_agg = df_envio_daily.groupby("DIA", as_index=False)["QTD"].sum().rename(columns={"QTD": "ENVIADAS"})
    
    df_receb_daily = df_receb_filtered.copy()
    df_receb_daily["DIA"] = pd.to_datetime(df_receb_daily["DIA"]).dt.date
    receb_daily_agg = df_receb_daily.groupby("DIA", as_index=False)["REAL CORTADO"].sum().rename(columns={"REAL CORTADO": "RECEBIDAS"})
    
    merged_daily = pd.merge(envio_daily_agg, receb_daily_agg, on="DIA", how="outer").fillna(0)
    merged_daily = merged_daily.sort_values("DIA").reset_index(drop=True)
    
    # Layout de Cartões KPI
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(
            f"""
            <div class="metric-card-exec">
                <div class="label">Peças Enviadas</div>
                <div class="value">{kpis['enviadas']:,}</div>
                <div class="sub">Total peças p/ produção</div>
            </div>
            """.replace(",", "."),
            unsafe_allow_html=True
        )
    with col2:
        st.markdown(
            f"""
            <div class="metric-card-exec" style="border-left-color: #059669;">
                <div class="label">Peças Recebidas</div>
                <div class="value">{kpis['recebidas']:,}</div>
                <div class="sub">Produção finalizada</div>
            </div>
            """.replace(",", "."),
            unsafe_allow_html=True
        )
    with col3:
        st.markdown(
            f"""
            <div class="metric-card-exec" style="border-left-color: #1E3A8A;">
                <div class="label">Minutos Enviados</div>
                <div class="value">{int(kpis['minutos_enviados']):,}</div>
                <div class="sub">Carga horária enviada</div>
            </div>
            """.replace(",", "."),
            unsafe_allow_html=True
        )
    with col4:
        st.markdown(
            f"""
            <div class="metric-card-exec" style="border-left-color: #D97706;">
                <div class="label">Minutos Recebidos</div>
                <div class="value">{int(kpis['minutos_recebidos']):,}</div>
                <div class="sub">Carga horária entregue</div>
            </div>
            """.replace(",", "."),
            unsafe_allow_html=True
        )
        
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Seção do Gráfico (Largura Cheia)
    st.markdown("#### 📈 Evolução Temporal Diária (Envio vs Recebimento)")
    if not merged_daily.empty:
        safe_st_echarts(_build_chart_comparison(merged_daily), height="400px")
    else:
        st.info("Sem dados suficientes para gerar o gráfico comparativo.")
        
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Seção da Tabela (Abaixo do Gráfico)
    st.markdown("#### 🏭 Resumo por Oficina")
    
    # Criação de tabela executiva em HTML (sem espaços de recuo para evitar bloco de código markdown)
    table_html = (
        '<div class="executive-table-wrapper">'
        '<table class="executive-table">'
        '<thead>'
        '<tr>'
        '<th>Oficina</th>'
        '<th class="num">Enviadas</th>'
        '<th class="num">Recebidas</th>'
        '<th class="num">Saldo</th>'
        '<th class="num">Retorno</th>'
        '</tr>'
        '</thead>'
        '<tbody>'
    )
    for _, row in summary_df.head(10).iterrows():
        saldo_val = f"{row['SALDO_PENDENTE']:,}".replace(",", ".")
        env_val = f"{row['ENVIADAS']:,}".replace(",", ".")
        rec_val = f"{row['RECEBIDAS']:,}".replace(",", ".")
        color_val = '#0E9F6E' if row['SALDO_PENDENTE'] >= 0 else '#B91C1C'
        
        table_html += (
            f'<tr>'
            f'<td><strong>{row["OFICINA"]}</strong></td>'
            f'<td class="num">{env_val}</td>'
            f'<td class="num">{rec_val}</td>'
            f'<td class="num" style="color:{color_val}; font-weight:bold;">{saldo_val}</td>'
            f'<td class="num"><strong>{row["TAXA_RETORNO"]:.1f}%</strong></td>'
            f'</tr>'
        )
        
    table_html += (
        '</tbody>'
        '</table>'
        '</div>'
        '<p style="font-size: 11px; color: #6B7A90; margin-top: 5px; margin-bottom: 15px; line-height: 1.4;">'
        '* <strong>Nota sobre consistência de saldo</strong>: Saldos negativos ou taxas de retorno superiores a 100% refletem a defasagem temporal de produção (peças que foram enviadas para a oficina em períodos anteriores ao selecionado e que retornaram/foram finalizadas dentro do período atual analisado).'
        '</p>'
    )
    st.markdown(table_html, unsafe_allow_html=True)
    
    # Botão de exportação
    excel_data = _to_excel(summary_df, df_envio_filtered, df_receb_filtered, d_ini, d_fim)
    st.download_button(
        label="📥 Exportar Relatório Consolidado (Excel)",
        data=excel_data,
        file_name=f"Relatorio_Executivo_{date.today().strftime('%d_%m_%Y')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
            
    st.divider()
    st.markdown(
        """
        <div style="text-align:center; font-size:12px; color:#A0C4B3; padding:12px;">
            Relatório de Gestão Executiva — APP Analise Oficina
        </div>
        """,
        unsafe_allow_html=True
    )
